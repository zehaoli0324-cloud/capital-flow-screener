"""
screener_core.py — 选股核心逻辑（Web版，无 input()，_pro 由外部注入）
"""
import os, sys, time, warnings, re, random, pathlib, io
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout, as_completed

warnings.filterwarnings("ignore")

import tushare as _ts
import requests
import requests.adapters
import urllib3
urllib3.disable_warnings()

from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# _pro 由 app.py 注入，此处初始化为 None
_pro = None

# ── 全局会话 ──────────────────────────────────────────────────
_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
       "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

_S = requests.Session()
_S.trust_env = False
_S.verify    = False
_retry = Retry(total=3, backoff_factor=1,
               status_forcelist=[429, 500, 502, 503, 504],
               allowed_methods=["HEAD", "GET", "OPTIONS"],
               raise_on_status=False)
_adapter = HTTPAdapter(max_retries=_retry, pool_connections=30, pool_maxsize=30)
_S.mount("https://", _adapter)
_S.mount("http://",  _adapter)
_S.headers.update({
    "User-Agent":      _UA,
    "Accept":          "text/html,application/xhtml+xml,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection":      "keep-alive",
})

_PROXY_URL = None  # 由 app.py 在启动时设置

def _get(url, params=None, headers=None, timeout=20, retries=3):
    h = dict(_S.headers)
    if headers:
        h.update(headers)
    DIRECT = {"http": _PROXY_URL, "https": _PROXY_URL} if _PROXY_URL else {"http": None, "https": None}
    for i in range(retries):
        try:
            r = _S.get(url, params=params, headers=h,
                       proxies=DIRECT, verify=False, timeout=timeout)
            r.raise_for_status()
            return r
        except Exception:
            if i == retries - 1:
                raise
            time.sleep(1 * (i + 1))

STAGE1_TOPN   = 50    # 量价筛选后进入第二阶段的股票数
LOOKBACK      = 10    # 资金流向回溯天数
KLINE_DAYS    = 90    # K线回溯天数
FF_TIMEOUT    = 12    # 资金流向获取超时秒数
FF_WORKERS    = 8     # 资金流向并发线程数

BANNER = """
╔══════════════════════════════════════════════════════╗
║  capital_flow_screener  v5  |  Tushare 稳定版        ║
║  步骤1: Tushare全量快照  →  步骤2: 量价评分          ║
║  →  步骤3: 资金流向深度评分(18分项)  →  综合输出     ║
╚══════════════════════════════════════════════════════╝
"""

# ══════════════════════════════════════════════════════════════
# 日期输入
# ══════════════════════════════════════════════════════════════

def input_target_date() -> datetime:
    raw = input("请输入分析日期（YYYYMMDD 或 YYYY-MM-DD，直接回车=今日）：").strip()
    if not raw:
        return datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    for fmt in ("%Y%m%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            pass
    print("  日期格式不识别，使用今日")
    return datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

# ══════════════════════════════════════════════════════════════
# 步骤 1：Tushare 全量行情快照
# ══════════════════════════════════════════════════════════════

def _ts_code_to_6(ts_code: str) -> str:
    """000001.SZ → 000001"""
    return ts_code.split(".")[0] if "." in ts_code else ts_code

def get_spot_data(target_date: datetime = None) -> pd.DataFrame:
    """
    用 Tushare pro.daily + pro.daily_basic 获取全市场行情快照。
    target_date 为 None 时取最近交易日（今日或最近一个交易日）。
    """
    print("    [Tushare] 获取全市场行情...")
    if target_date is None:
        trade_date = datetime.today().strftime("%Y%m%d")
    else:
        trade_date = target_date.strftime("%Y%m%d")

    # 若当天非交易日，自动回退到最近一个交易日
    for attempt in range(10):
        try:
            df = _pro.daily(trade_date=trade_date,
                            fields="ts_code,open,high,low,close,vol,amount,pct_chg,turnover_rate")
            if df is not None and len(df) > 100:
                break
        except Exception as e:
            print(f"    Tushare daily 失败({attempt+1})：{e}")
        trade_date = (datetime.strptime(trade_date, "%Y%m%d") - timedelta(days=1)).strftime("%Y%m%d")
        time.sleep(0.5)
    else:
        raise RuntimeError("Tushare 无法获取全市场行情，请检查Token和积分")

    # 同时拉每日指标（市值、量比）
    # ⚠️ daily 已含 turnover_rate，daily_basic 也含 turnover_rate
    #    避免列名冲突：从 daily_basic 只取 volume_ratio / circ_mv，换手率直接用 daily 自带的
    try:
        basic = _pro.daily_basic(trade_date=trade_date,
                                 fields="ts_code,volume_ratio,circ_mv")
        if basic is not None and len(basic) > 0:
            merge_cols = ["ts_code"]
            if "volume_ratio" in basic.columns: merge_cols.append("volume_ratio")
            if "circ_mv"      in basic.columns: merge_cols.append("circ_mv")
            df = df.merge(basic[merge_cols], on="ts_code", how="left")
    except Exception as e:
        print(f"    daily_basic 获取失败（非致命）：{e}")

    if "volume_ratio" not in df.columns: df["volume_ratio"] = np.nan
    if "circ_mv"      not in df.columns: df["circ_mv"]      = np.nan

    df["code"]      = df["ts_code"].apply(_ts_code_to_6)
    df["price"]     = pd.to_numeric(df["close"],        errors="coerce")
    df["pct_chg"]   = pd.to_numeric(df["pct_chg"],      errors="coerce")
    df["volume"]    = pd.to_numeric(df["vol"],           errors="coerce")
    # turnover_rate 直接来自 daily 接口（已请求该字段），无需 fallback
    df["turnover"]  = pd.to_numeric(df["turnover_rate"], errors="coerce") \
                      if "turnover_rate" in df.columns else np.nan
    df["vol_ratio"] = pd.to_numeric(df.get("volume_ratio",  np.nan), errors="coerce")
    df["circ_cap_yi"] = pd.to_numeric(df.get("circ_mv", np.nan), errors="coerce") / 10000  # 万元→亿元

    # 取名称（stock_basic，一次性缓存）
    try:
        sb = _pro.stock_basic(fields="ts_code,name")
        if sb is not None and len(sb) > 0:
            name_map = dict(zip(sb["ts_code"].apply(_ts_code_to_6), sb["name"]))
            df["name"] = df["code"].map(name_map).fillna(df["code"])
        else:
            df["name"] = df["code"]
    except Exception:
        df["name"] = df["code"]

    # 过滤：只保留主板 A 股（沪深主板），排除科创板/创业板/北交所/ST
    df = df[df["code"].str.match(r"^(0|6)\d{5}$")].copy()
    df = df[~df["name"].str.contains("ST|退", na=False)]
    df.reset_index(drop=True, inplace=True)

    actual_date = trade_date
    print(f"    [Tushare] 成功，共 {len(df)} 只（{actual_date}）")
    return df, actual_date

def get_hs300_change(trade_date: str = None) -> float:
    """沪深300当日涨跌幅，用 Tushare index_daily 获取"""
    try:
        today = trade_date or datetime.today().strftime("%Y%m%d")
        start = (datetime.strptime(today, "%Y%m%d") - timedelta(days=10)).strftime("%Y%m%d")
        df = _pro.index_daily(ts_code="399300.SZ", start_date=start, end_date=today,
                              fields="trade_date,pct_chg")
        if df is not None and len(df) > 0:
            df.sort_values("trade_date", inplace=True)
            return float(df.iloc[-1]["pct_chg"])
    except Exception as e:
        print(f"    沪深300获取失败：{e}")
    return 0.0

# ══════════════════════════════════════════════════════════════
# 个股 K 线（用于量价评分历史K线 + 资金流向二期指标）
# ══════════════════════════════════════════════════════════════

def fetch_kline(code: str, days: int = KLINE_DAYS, end_date: str = None):
    """
    用 Tushare pro_bar 获取前复权日K线。
    返回标准化 DataFrame（date, open, close, high, low, volume, turnover）
    turnover 来源：优先 pro_bar factors=[tor]，若全为 nan 则回退 pro.daily turnover_rate
    """
    suffix   = ".SH" if code.startswith(("6", "5")) else ".SZ"
    ts_code  = code + suffix
    end_dt   = end_date or datetime.today().strftime("%Y%m%d")
    start_dt = (datetime.strptime(end_dt, "%Y%m%d") - timedelta(days=days + 30)).strftime("%Y%m%d")
    try:
        import tushare as _ts
        df = _ts.pro_bar(ts_code=ts_code, adj="qfq",
                        start_date=start_dt, end_date=end_dt,
                        factors=["tor"], freq="D")
        if df is None or not isinstance(df, pd.DataFrame) or len(df) < 10:
            return None

        # ── 列名标准化（pro_bar 偶尔多返回 adj_factor 等列，只取需要的）──
        col_map = {"trade_date": "date", "vol": "volume", "tor": "turnover"}
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

        # 只保留必要列，防止无关列引发后续错误
        keep = [c for c in ["date", "open", "close", "high", "low", "volume", "turnover"]
                if c in df.columns]
        df = df[keep].copy()

        if "date" not in df.columns:
            return None
        if "turnover" not in df.columns:
            df["turnover"] = np.nan

        for c in ["open", "close", "high", "low", "volume", "turnover"]:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")

        df["date"] = pd.to_datetime(df["date"].astype(str), format="%Y%m%d", errors="coerce")
        df.dropna(subset=["date", "close", "volume"], inplace=True)
        df.sort_values("date", ascending=True, inplace=True)
        df.reset_index(drop=True, inplace=True)
        df = df.tail(days).copy()

        # ── 换手率 nan 回退：tor 全为空时用 pro.daily turnover_rate 补充 ──
        if df["turnover"].isna().all():
            try:
                d2 = _pro.daily(ts_code=ts_code, start_date=start_dt, end_date=end_dt,
                                fields="trade_date,turnover_rate")
                if d2 is not None and isinstance(d2, pd.DataFrame) and len(d2) > 0:
                    d2["date"] = pd.to_datetime(d2["trade_date"].astype(str),
                                                format="%Y%m%d", errors="coerce")
                    d2["turnover_rate"] = pd.to_numeric(d2["turnover_rate"], errors="coerce")
                    to_map = dict(zip(d2["date"], d2["turnover_rate"]))
                    df["turnover"] = df["date"].map(to_map)
            except Exception:
                pass

        return df if len(df) >= 5 else None
    except Exception:
        return None

# ══════════════════════════════════════════════════════════════
# 步骤 2：量价评分（全本地，无网络请求）
# ══════════════════════════════════════════════════════════════

def screen_stage1(df: pd.DataFrame, hs300_chg: float,
                  actual_date: str, ff_workers: int = 4) -> list[dict]:
    """
    对快照 DataFrame 做量价评分，返回通过的股票记录列表（含 stage1_score）。
    K线历史数据通过 fetch_kline 并发获取（用于均线/量比等条件）。
    """
    if hs300_chg <= -2.0:
        mkt_mode, pct_lo, pct_hi, vol_thresh, to_lo, score_pass = "弱市",  0.5, 9.5, 1.2, 2.0, 3
    elif hs300_chg <= 0:
        mkt_mode, pct_lo, pct_hi, vol_thresh, to_lo, score_pass = "震荡",  1.0, 9.5, 1.5, 2.0, 3
    else:
        mkt_mode, pct_lo, pct_hi, vol_thresh, to_lo, score_pass = "强市",  2.0, 9.5, 2.0, 3.0, 4

    # 强市额外收紧：涨幅下限再提高（跑赢大盘至少1%），防止候选池过大
    if hs300_chg > 0:
        pct_lo = max(pct_lo, hs300_chg + 1.0)

    print(f"  [评分] 大盘【{mkt_mode}】→ 涨幅预筛 {pct_lo:.1f}%~{pct_hi}%，量比≥{vol_thresh}，换手≥{to_lo}%")

    # 一次性预筛（涨幅/量比/换手），并按综合热度排序后截取前300
    pre_filtered = []
    for _, row in df.iterrows():
        pct_chg   = row.get("pct_chg",  np.nan)
        vol_ratio = row.get("vol_ratio", np.nan)
        turnover  = row.get("turnover",  np.nan)
        if pd.isna(pct_chg) or not (pct_lo <= pct_chg <= pct_hi):
            continue
        if pd.notna(vol_ratio) and vol_ratio < vol_thresh:
            continue
        if pd.notna(turnover) and turnover < to_lo:
            continue
        pre_filtered.append(row.to_dict())

    # 按"涨幅×量比"热度降序，截取前300只再拉K线，避免强市过慢
    PRE_CAP = 300
    if len(pre_filtered) > PRE_CAP:
        pre_filtered.sort(
            key=lambda r: float(r.get("pct_chg", 0) or 0) * float(r.get("vol_ratio", 1) or 1),
            reverse=True
        )
        pre_filtered = pre_filtered[:PRE_CAP]
        print(f"  [评分] 预筛结果过多，已按热度截取前 {PRE_CAP} 只")

    print(f"  [评分] 涨幅/量比/换手预筛：{len(pre_filtered)} 只，并发拉取K线...")

    # 并发拉 K 线（带实时进度条）
    # 进度条用 \r 原地刷新；fetch_kline 内所有异常已被捕获不会打印，
    # 但 Tushare SDK 内部偶尔会有 warnings，用 warnings.catch_warnings 屏蔽
    import warnings as _warnings
    kline_map  = {}
    total_k    = len(pre_filtered)
    done_k     = 0
    ok_k       = 0
    t_k_start  = time.time()
    _k_lock    = __import__("threading").Lock()

    def _fetch_kline_tracked(code):
        nonlocal done_k, ok_k
        with _warnings.catch_warnings():
            _warnings.simplefilter("ignore")
            result = fetch_kline(code, KLINE_DAYS, actual_date)
        with _k_lock:
            done_k += 1
            if result is not None:
                ok_k += 1
            elapsed = time.time() - t_k_start
            eta     = (elapsed / done_k * (total_k - done_k)) if done_k else 0
            bar_w   = 30
            fill    = int(bar_w * done_k / total_k)
            bar_str = "█" * fill + "░" * (bar_w - fill)
            # 用固定宽度字段避免残影
            line = (f"  [{bar_str}] {done_k:>3}/{total_k}  "
                    f"成功:{ok_k:>3}  失败:{done_k-ok_k:>2}  "
                    f"已用:{int(elapsed):>3}s  ETA:{int(eta):>3}s")
            print(f"\r{line:<90}", end="", flush=True)
        return code, result

    with ThreadPoolExecutor(max_workers=ff_workers) as pool:
        futs = {pool.submit(_fetch_kline_tracked, row["code"]): row["code"]
                for row in pre_filtered}
        for fut in as_completed(futs):
            code, kdf = fut.result()
            kline_map[code] = kdf

    # 进度条结束：打印完整的最终状态行再换行
    elapsed_total = time.time() - t_k_start
    bar_str = "█" * 30
    print(f"\r  [{bar_str}] {total_k:>3}/{total_k}  "
          f"成功:{ok_k:>3}  失败:{total_k-ok_k:>2}  "
          f"已用:{int(elapsed_total):>3}s  ETA:  0s")
    print(f"  [评分] K线拉取完成：{ok_k}/{total_k} 只有效，耗时 {elapsed_total:.1f}s")

    candidates = []
    for row in pre_filtered:
        code      = str(row.get("code", "")).zfill(6)
        hist      = kline_map.get(code)
        sc, hits  = _score_one(row, hs300_chg, hist)
        if sc < score_pass:
            continue
        candidates.append({
            "code":         code,
            "name":         str(row.get("name", code)),
            "price":        float(row.get("price", 0) or 0),
            "pct_chg":      float(row.get("pct_chg", 0)),
            "vol_ratio":    float(row.get("vol_ratio", 0) or 0),
            "turnover":     float(row["turnover"]) if pd.notna(row.get("turnover")) else np.nan,
            "circ_cap_yi":  float(row.get("circ_cap_yi", 0) or 0),
            "stage1_score": sc,
            "stage1_hits":  hits,
            "_hist":        hist,   # 传递给后续二期评分
        })

    candidates.sort(key=lambda x: (x["stage1_score"], x["pct_chg"]), reverse=True)
    return candidates


def _score_one(row: dict, hs300_chg: float, hist=None) -> tuple[int, list]:
    """对单只股票打量价分，返回 (score, hits列表)"""
    pct_chg   = row.get("pct_chg",  np.nan)
    vol_ratio = row.get("vol_ratio", np.nan)
    turnover  = row.get("turnover",  np.nan)
    hits = []

    # ① 涨幅 3-8%
    if pd.notna(pct_chg) and 3.0 <= pct_chg <= 8.0:
        hits.append("①涨幅3-8%")
    # ② 换手 5-15%
    if pd.notna(turnover) and 5.0 <= turnover <= 15.0:
        hits.append("②换手5-15%")

    if hist is not None and len(hist) >= 6:
        close = hist["close"]
        vol   = hist["volume"]
        open_ = hist["open"]
        high  = hist["high"]

        # ③ 放量≥1.5x（比5日均量）
        if len(vol) >= 6:
            avg5 = vol.iloc[-6:-1].mean()
            if pd.notna(avg5) and avg5 > 0 and vol.iloc[-1] >= avg5 * 1.5:
                hits.append("③放量≥1.5x")
        # ④ 超额跑赢大盘+1%
        if pd.notna(pct_chg) and pct_chg > hs300_chg + 1.0:
            hits.append("④超额+1%")
        # ⑤ 量能 TOP 10%（近60日）
        if len(vol) >= 10:
            rank_pct = (vol.iloc[-min(60, len(vol)):] < vol.iloc[-1]).mean()
            if rank_pct >= 0.90:
                hits.append("⑤量TOP10%")
        # ⑥ 价格站上均线
        if len(close) >= 55:
            ma13 = close.rolling(13).mean().iloc[-1]
            ma34 = close.rolling(34).mean().iloc[-1]
            ma55 = close.rolling(55).mean().iloc[-1]
            if pd.notna(ma13) and pd.notna(ma34) and pd.notna(ma55) and close.iloc[-1] > ma13 > ma34 > ma55:
                hits.append("⑥价格>MA13>MA34>MA55")
        elif len(close) >= 34:
            ma13 = close.rolling(13).mean().iloc[-1]
            ma34 = close.rolling(34).mean().iloc[-1]
            if pd.notna(ma13) and pd.notna(ma34) and close.iloc[-1] > ma13 > ma34:
                hits.append("⑥价格>MA13>MA34")
        elif len(close) >= 13:
            ma13 = close.rolling(13).mean().iloc[-1]
            if pd.notna(ma13) and close.iloc[-1] > ma13:
                hits.append("⑥价格>MA13")
        # ⑦ 近5日放量阳线≥2根
        if len(vol) >= 10:
            avg_ref = vol.iloc[-10:-5].mean()
            yang = sum(
                1 for c, o, v in zip(close.iloc[-5:].values,
                                     open_.iloc[-5:].values,
                                     vol.iloc[-5:].values)
                if c > o and pd.notna(avg_ref) and avg_ref > 0 and v > avg_ref
            )
            if yang >= 2:
                hits.append(f"⑦近5日{yang}根放量阳线")
        # ⑧ 距近期高点回调<15%
        if len(high) >= 10:
            h_ref = high.iloc[-min(60, len(high)):-1].max()
            draw  = (h_ref - close.iloc[-1]) / h_ref * 100 if h_ref > 0 else 999
            if draw < 15.0:
                hits.append(f"⑧距高点回调{draw:.1f}%")

    return len(hits), hits

# ══════════════════════════════════════════════════════════════
# 步骤 3：资金流向获取（Tushare优先 七级数据源）
# ══════════════════════════════════════════════════════════════

def _tofloat(v):
    try: return float(v)
    except: return np.nan

def _build_ff_df(rows, yuan_unit=False):
    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df.dropna(subset=["date"], inplace=True)
    for col in ["超大单净额", "大单净额", "中单净额", "小单净额", "超大单净占比", "收盘价", "涨跌幅"]:
        df[col] = pd.to_numeric(df.get(col, np.nan), errors="coerce") if col in df.columns else np.nan
    if not yuan_unit:
        for col in ["超大单净额", "大单净额", "中单净额", "小单净额"]:
            if col in df.columns:
                df[col] = df[col] / 10000
    df.sort_values("date", ascending=False, inplace=True)
    df.reset_index(drop=True, inplace=True)
    df = df.head(LOOKBACK)
    if df["超大单净额"].abs().sum() < 0.01 or df["超大单净额"].isna().all():
        return None
    return df

def _ff_tushare(code: str, end_date: str = None):
    """
    Tushare moneyflow 接口（走 Tushare 自有服务器，最稳定）
    """
    suffix  = ".SH" if code.startswith(("6", "5")) else ".SZ"
    ts_code = code + suffix
    end_dt  = end_date or datetime.today().strftime("%Y%m%d")
    start_dt = (datetime.strptime(end_dt, "%Y%m%d") - timedelta(days=LOOKBACK + 10)).strftime("%Y%m%d")
    try:
        df = _pro.moneyflow(
            ts_code=ts_code,
            start_date=start_dt,
            end_date=end_dt,
            fields="trade_date,buy_elg_amount,sell_elg_amount,"
                   "buy_lg_amount,sell_lg_amount,"
                   "buy_md_amount,sell_md_amount,"
                   "buy_sm_amount,sell_sm_amount,"
                   "net_mf_amount,close,pct_change"
        )
        if df is None or len(df) == 0:
            return None
        df.sort_values("trade_date", ascending=False, inplace=True)
        df.reset_index(drop=True, inplace=True)

        rows = []
        for _, row in df.iterrows():
            def _v(col):
                try:
                    v = row.get(col)
                    return float(v) if v is not None and str(v) not in ("nan", "") else 0.0
                except: return 0.0
            super_net = _v("buy_elg_amount") - _v("sell_elg_amount")
            large_net = _v("buy_lg_amount")  - _v("sell_lg_amount")
            mid_net   = _v("buy_md_amount")  - _v("sell_md_amount")
            small_net = _v("buy_sm_amount")  - _v("sell_sm_amount")
            total_vol = abs(super_net) + abs(large_net) + abs(mid_net) + abs(small_net)
            super_pct = (super_net / total_vol * 100) if total_vol > 0 else 0.0
            rows.append({
                "date":        row["trade_date"],
                "超大单净额":  super_net,
                "大单净额":    large_net,
                "中单净额":    mid_net,
                "小单净额":    small_net,
                "超大单净占比": super_pct,
                "收盘价":      _v("close"),
                "涨跌幅":      _v("pct_change"),
            })
        if not rows:
            return None
        return _build_ff_df(rows, yuan_unit=True)
    except Exception:
        return None

def _ff_em_a(code: str, **kwargs):
    """东方财富 push2his 历史资金流"""
    market = "1" if code.startswith(("6", "5")) else "0"
    try:
        r = _get("https://push2his.eastmoney.com/api/qt/stock/fflow/daykline/get",
                 params={"lmt": "0", "klt": "101", "secid": f"{market}.{code}",
                         "fields1": "f1,f2,f3,f7",
                         "fields2": "f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61,f62,f63,f64,f65",
                         "ut": "b2884a393a59ad64002292a3e90d46a5",
                         "_": int(time.time() * 1000)},
                 headers={"Referer": "https://data.eastmoney.com/zjlx/detail.html",
                          "Origin": "https://data.eastmoney.com"})
        return _parse_em_ff(r.json().get("data", {}).get("klines", []))
    except: return None

def _ff_em_b(code: str, **kwargs):
    """东方财富 push2 备用 ut"""
    market = "1" if code.startswith(("6", "5")) else "0"
    try:
        r = _get("https://push2.eastmoney.com/api/qt/stock/fflow/daykline/get",
                 params={"lmt": "0", "klt": "101", "secid": f"{market}.{code}",
                         "fields1": "f1,f2,f3,f7",
                         "fields2": "f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61,f62,f63,f64,f65",
                         "ut": "bd1d9ddb04089700cf9c27f6f7426281",
                         "_": int(time.time() * 1000)},
                 headers={"Referer": "https://quote.eastmoney.com/"})
        return _parse_em_ff(r.json().get("data", {}).get("klines", []))
    except: return None

def _ff_em_c(code: str, **kwargs):
    """东方财富 datacenter-web 数据中心接口"""
    try:
        r = _get("https://datacenter-web.eastmoney.com/api/data/v1/get",
                 params={"sortColumns": "TRADE_DATE", "sortTypes": "-1",
                         "pageSize": str(LOOKBACK + 2), "pageNumber": "1",
                         "reportName": "RPT_CUSTOM_STOCK_CAPITAL",
                         "columns": "ALL",
                         "filter": f'(SECURITY_CODE="{code}")',
                         "token": "894050c76af8597a853f5b408b759f5d"},
                 headers={"Referer": "https://data.eastmoney.com/zjlx/detail.html",
                          "Origin": "https://data.eastmoney.com"}, timeout=10)
        data = r.json().get("result", {})
        if not data: return None
        items = data.get("data", [])
        if not items: return None
        rows = [{"date": it.get("TRADE_DATE", ""),
                 "超大单净额":  _tofloat(it.get("SUPER_LARGE_NET_AMOUNT")),
                 "大单净额":    _tofloat(it.get("LARGE_NET_AMOUNT")),
                 "中单净额":    _tofloat(it.get("MEDIUM_NET_AMOUNT")),
                 "小单净额":    _tofloat(it.get("SMALL_NET_AMOUNT")),
                 "超大单净占比": _tofloat(it.get("SUPER_LARGE_NET_RATIO")),
                 "收盘价":      _tofloat(it.get("CLOSE_PRICE")),
                 "涨跌幅":      _tofloat(it.get("CHANGE_RATE"))} for it in items]
        return _build_ff_df(rows, yuan_unit=True) if rows else None
    except: return None

def _ff_em_d(code: str, **kwargs):
    """东方财富 emappdata WAP 端接口"""
    market = "1" if code.startswith(("6", "5")) else "0"
    secid  = f"{market}.{code}"
    try:
        r = _get("https://emappdata.eastmoney.com/stockanalysis/real/getAllHisSecurityMarketCapFlow",
                 params={"appVersion": "10.3.1.1", "deviceid": "wap", "plat": "Wap",
                         "product": "EFund", "version": "1.0.0", "securityCode": secid},
                 headers={"Referer": "https://wap.eastmoney.com/",
                          "Origin": "https://wap.eastmoney.com"}, timeout=10)
        items = r.json().get("data") or []
        if not items: return None
        rows = [{"date":          it.get("date", ""),
                 "超大单净额":    _tofloat(it.get("superLargeNetAmount")),
                 "大单净额":      _tofloat(it.get("largeNetAmount")),
                 "中单净额":      _tofloat(it.get("middleNetAmount")),
                 "小单净额":      _tofloat(it.get("smallNetAmount")),
                 "超大单净占比":  _tofloat(it.get("superLargeNetRatio")),
                 "收盘价":        _tofloat(it.get("closePrice")),
                 "涨跌幅":        _tofloat(it.get("changeRatio"))} for it in items]
        return _build_ff_df(rows, yuan_unit=True) if rows else None
    except: return None

def _ff_sina(code: str, **kwargs):
    """新浪财经个股历史资金流接口"""
    prefix = "sh" if code.startswith(("6", "5")) else "sz"
    try:
        r = _get(
            "https://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/MoneyFlow.ssi_ssfx_flzjlx",
            params={"daima": f"{prefix}{code}", "page": "1",
                    "num": str(LOOKBACK + 2), "sort": "opendate", "asc": "0", "type": "0"},
            headers={"Referer": "https://finance.sina.com.cn/"}, timeout=10)
        import json as _json
        text = r.text.strip()
        if not text or text == "null": return None
        items = _json.loads(text)
        if not items or not isinstance(items, list): return None
        rows = []
        for it in items:
            try:
                rows.append({"date":          it.get("opendate", ""),
                             "超大单净额":    _tofloat(it.get("r0_net", it.get("netamount"))),
                             "大单净额":      _tofloat(it.get("r1_net")),
                             "中单净额":      _tofloat(it.get("r2_net")),
                             "小单净额":      _tofloat(it.get("r3_net")),
                             "超大单净占比":  _tofloat(it.get("r0_ratio")),
                             "收盘价":        _tofloat(it.get("closeprice")),
                             "涨跌幅":        _tofloat(it.get("changeratio"))})
            except: continue
        return _build_ff_df(rows, yuan_unit=True) if rows else None
    except: return None

def _ff_qq(code: str, **kwargs):
    """腾讯行情资金流兜底（仅当日单条）"""
    try:
        prefix = "sh" if code.startswith(("6", "5")) else "sz"
        r = _get(f"https://qt.gtimg.cn/q=ff_{prefix}{code}",
                 headers={"Referer": "https://finance.qq.com"}, timeout=8)
        try:    text = r.content.decode("gbk", errors="replace")
        except: text = r.text
        raw = text.split('"'[0])[1] if '"'[0] in text else ""
        if not raw or raw == "-": return None
        f = raw.split("~")
        if len(f) < 10: return None
        def _v(i):
            try: return float(f[i]) / 10000 if f[i].strip() not in ("", "-", "--") else np.nan
            except: return np.nan
        si, so = _v(1), _v(2)
        li, lo = _v(3), _v(4)
        mi, mo = _v(5), _v(6)
        xi, xo = _v(7), _v(8)
        if np.isnan(si) and np.isnan(so): return None
        rec = {"date":        pd.Timestamp.today(),
               "超大单净额": (si or 0) - (so or 0), "大单净额":  (li or 0) - (lo or 0),
               "中单净额":   (mi or 0) - (mo or 0), "小单净额":  (xi or 0) - (xo or 0),
               "超大单净占比": np.nan, "收盘价": np.nan, "涨跌幅": np.nan}
        return _build_ff_df([rec])
    except: return None

def _parse_em_ff(klines):
    if not klines: return None
    rows = []
    for item in klines:
        f = item.split(",")
        if len(f) < 13: continue
        rows.append({"date":          f[0],
                     "超大单净额":    _tofloat(f[5]),
                     "大单净额":      _tofloat(f[4]),
                     "中单净额":      _tofloat(f[3]),
                     "小单净额":      _tofloat(f[2]),
                     "超大单净占比":  _tofloat(f[10]),
                     "收盘价":        _tofloat(f[11]),
                     "涨跌幅":        _tofloat(f[12])})
    return _build_ff_df(rows) if rows else None

def fetch_fund_flow(code: str, end_date: str = None):
    """单只股票资金流向（七级数据源，Tushare 优先）"""
    with ThreadPoolExecutor(max_workers=1) as ex:
        fut = ex.submit(_fetch_ff_inner, code, end_date)
        try:
            return fut.result(timeout=FF_TIMEOUT)
        except FuturesTimeout:
            return None
        except: return None

def _fetch_ff_inner(code: str, end_date: str = None):
    for fn in [_ff_tushare, _ff_em_a, _ff_em_b, _ff_em_c, _ff_em_d, _ff_sina, _ff_qq]:
        try:
            df = fn(code, end_date=end_date)
            if df is not None: return df
        except: pass
    return None

def fetch_fund_flows(candidates: list[dict], actual_date: str) -> dict:
    """并发批量拉取候选股资金流向"""
    print(f"\n  [资金流向] 并发拉取 {len(candidates)} 只（{FF_WORKERS} 线程，"
          f"单只超时 {FF_TIMEOUT}s）...")
    results = {}

    def _worker(rec):
        code = rec["code"]
        try:
            df = fetch_fund_flow(code, end_date=actual_date)
        except Exception:
            df = None
        return code, df

    total_ff   = len(candidates)
    ff_t_start = time.time()

    with ThreadPoolExecutor(max_workers=FF_WORKERS) as pool:
        futures = {pool.submit(_worker, rec): rec for rec in candidates}
        done = ok = 0
        for fut in as_completed(futures):
            code, df = fut.result()
            results[code] = df
            done += 1
            if df is not None: ok += 1
            elapsed_ff = time.time() - ff_t_start
            eta_ff     = (elapsed_ff / done * (total_ff - done)) if done else 0
            bar_fill   = int(30 * done / total_ff)
            bar_str    = "█" * bar_fill + "░" * (30 - bar_fill)
            print(f"\r  [{bar_str}] {done}/{total_ff}  "
                  f"成功:{ok}  失败:{done-ok}  "
                  f"剩余:{int(eta_ff):>3}s",
                  end="", flush=True)
            time.sleep(0.1)

    print()
    print(f"  [资金流向] 完成：{ok}/{total_ff} 只获取到数据，"
          f"耗时 {time.time()-ff_t_start:.1f}s")
    return results

# ══════════════════════════════════════════════════════════════
# 二期 K 线衍生指标（来自 fund_flow_scorer_v6）
# ══════════════════════════════════════════════════════════════

def calc_cmf(kdf, n=14) -> float:
    if kdf is None or len(kdf) < n: return np.nan
    df = kdf.tail(n).copy()
    hl = (df["high"] - df["low"]).replace(0, np.nan)
    mfm = ((df["close"] - df["low"]) - (df["high"] - df["close"])) / hl
    mfv = mfm * df["volume"]
    vol_sum = df["volume"].sum()
    return float(mfv.sum() / vol_sum) if vol_sum > 0 else 0.0

def calc_obv_slope(kdf, n=10) -> float:
    if kdf is None or len(kdf) < n + 1: return np.nan
    df = kdf.copy()
    obv = [0.0]
    for i in range(1, len(df)):
        if df["close"].iloc[i] > df["close"].iloc[i - 1]:
            obv.append(obv[-1] + df["volume"].iloc[i])
        elif df["close"].iloc[i] < df["close"].iloc[i - 1]:
            obv.append(obv[-1] - df["volume"].iloc[i])
        else:
            obv.append(obv[-1])
    seg = np.array(obv[-n:])
    slope = float(np.polyfit(range(n), seg, 1)[0])
    mean_abs = float(np.nanmean(np.abs(seg))) + 1
    return slope / mean_abs

def calc_mfi(kdf, n=14) -> float:
    if kdf is None or len(kdf) < n + 1: return np.nan
    df = kdf.tail(n + 1).copy().reset_index(drop=True)
    tp = (df["high"] + df["low"] + df["close"]) / 3
    pos = neg = 0.0
    for i in range(1, len(df)):
        mf = tp.iloc[i] * df["volume"].iloc[i]
        if tp.iloc[i] > tp.iloc[i - 1]: pos += mf
        elif tp.iloc[i] < tp.iloc[i - 1]: neg += mf
    return float(100 - 100 / (1 + pos / neg)) if neg > 0 else 100.0

def calc_vwap_dev(kdf, n=10) -> float:
    if kdf is None or len(kdf) < n: return np.nan
    df = kdf.tail(n).copy()
    vwap = (df["close"] * df["volume"]).sum() / df["volume"].sum()
    last = float(df["close"].iloc[-1])
    return float((last - vwap) / vwap * 100) if vwap > 0 else 0.0

def calc_vol_pattern(kdf, n=10) -> dict:
    if kdf is None or len(kdf) < 5:
        return {"shrink_up": 0, "expand_down": 0, "avg_to": np.nan, "lock": 3}
    df = kdf.copy()
    avg_vol = df["volume"].mean()
    recent  = df.tail(n)
    shrink_up = expand_down = 0
    for i in range(1, len(recent)):
        row = recent.iloc[i]; prev = recent.iloc[i - 1]
        is_up   = row["close"] > prev["close"]
        is_down = row["close"] < prev["close"]
        if is_up   and row["volume"] < avg_vol * 0.85: shrink_up   += 1
        if is_down and row["volume"] > avg_vol * 1.30: expand_down += 1
    to_col = "turnover" if "turnover" in df.columns else None
    avg_to = float(df[to_col].tail(10).mean()) if to_col else np.nan
    cum_to = float(df[to_col].tail(20).sum())  if to_col else np.nan
    if not np.isnan(cum_to):
        lock = 10 if cum_to < 15 else (7 if cum_to < 30 else (4 if cum_to < 50 else (2 if cum_to < 80 else 0)))
    else:
        lock = 3
    return {"shrink_up": shrink_up, "expand_down": expand_down, "avg_to": avg_to, "lock": lock}

def calc_high_risk(kdf) -> dict:
    if kdf is None or len(kdf) < 10:
        return {"dist_h": np.nan, "dist_l": np.nan}
    df   = kdf.tail(60)
    last = float(df["close"].iloc[-1])
    h60  = float(df["high"].max()); l60 = float(df["low"].min())
    return {"dist_h": (h60 - last) / h60 * 100 if h60 > 0 else np.nan,
            "dist_l": (last - l60) / l60 * 100 if l60 > 0 else np.nan}

# ══════════════════════════════════════════════════════════════
# 资金流向综合评分引擎 v3（移植自 fund_flow_scorer_v6）
# ══════════════════════════════════════════════════════════════

def score_fund_flow(code: str, ff, kdf, circ_cap_yi) -> dict:
    """
    ff          : 资金流向 DataFrame
    kdf         : 日K线 DataFrame（升序）
    circ_cap_yi : 流通市值（亿元）
    返回评分结果 dict，总分 0-100
    """
    base = {
        "s1": 0, "s2": 0, "s3": 0, "s4": 0, "s5": 0, "s6": 0,
        "n9_corr": 0, "n12_eff": 0, "n13_cv": 0, "n14_ratio": 0,
        "n5_cmf": 0, "n7_obv": 0, "n6_mfi": 0, "n10_vwap": 0,
        "n11_vol": 0, "n8_lock": 0, "r4_high": 0, "r5_turn": 0,
        "penalty_coef": 1.0, "bonus": 0, "total": 0,
        "direction": "→ 观望", "pattern": "无数据",
        "evidence": [], "risk": [],
        "consec_in": 0, "consec_out": 0, "recent_pct": 0.0,
        "net_3d": 0.0, "net_5d": 0.0, "slope_norm": 0.0,
        "large_corr": 0.0, "dark_pool_days": 0, "dump_days": 0,
        "flow_cv": 0.0, "st_ratio": 1.0,
        "cmf_val": np.nan, "obv_slope": np.nan, "mfi_val": np.nan,
        "vwap_dev": np.nan, "dist_high": np.nan, "avg_turnover": np.nan,
        "has_kline": kdf is not None,
        "has_ff":    ff is not None,
    }
    ev, rk = base["evidence"], base["risk"]

    if ff is None or len(ff) == 0:
        rk.append("未能获取资金流向数据")
        base["total"]     = 0
        base["direction"] = "→ 观望（无资金数据）"
        base["pattern"]   = "无资金数据"
        return base

    n         = len(ff)
    super_net = ff["超大单净额"].fillna(0).values
    large_net = ff["大单净额"].fillna(0).values
    mid_net   = ff["中单净额"].fillna(0).values
    small_net = ff["小单净额"].fillna(0).values
    super_pct = ff["超大单净占比"].fillna(0).values
    pct_chg   = ff["涨跌幅"].fillna(0).values

    # S1 持续性 /25
    ci = co = 0
    for v in super_net:
        if v > 0 and co == 0: ci += 1
        elif v < 0 and ci == 0: co += 1
        else: break
    net_in_days = int((super_net > 0).sum())
    if   ci >= 5: s1 = 25; ev.append(f"超大单连续{ci}日净流入，建仓意愿极强")
    elif ci == 4: s1 = 22; ev.append(f"超大单连续{ci}日净流入")
    elif ci == 3: s1 = 18; ev.append("超大单连续3日净流入，节奏稳定")
    elif ci == 2: s1 = 13; ev.append("超大单连续2日净流入")
    elif ci == 1: s1 =  7; ev.append("超大单近1日净流入，持续性待观察")
    elif co >= 4: s1 =  0; rk.append(f"超大单连续{co}日净流出，主力持续离场")
    elif co >= 3: s1 =  2; rk.append(f"超大单连续{co}日净流出")
    elif (super_net < 0).sum() > net_in_days: s1 = 4; rk.append("净流出天数多于净流入")
    else: s1 = 5
    base["s1"] = s1

    # S2 强度 /20
    recent_pct = float(np.nanmean(super_pct[:min(3, n)]))
    net_3d     = float(np.nansum(super_net[:min(3, n)]))
    net_5d     = float(np.nansum(super_net[:min(5, n)]))
    if   recent_pct >=  6.0: s2 = 20; ev.append(f"超大单净占比{recent_pct:.1f}%，机构介入极强")
    elif recent_pct >=  4.0: s2 = 17; ev.append(f"超大单净占比{recent_pct:.1f}%，介入明显")
    elif recent_pct >=  2.0: s2 = 13; ev.append(f"超大单净占比{recent_pct:.1f}%，有效介入")
    elif recent_pct >=  0.8: s2 =  8; ev.append(f"超大单净占比{recent_pct:.1f}%，介入力度偏弱")
    elif recent_pct <= -4.0: s2 =  0; rk.append(f"超大单净流出{abs(recent_pct):.1f}%，出货力度大")
    elif recent_pct  <  0:   s2 =  2; rk.append(f"超大单小幅净流出({abs(recent_pct):.1f}%)")
    else:                    s2 =  4
    base["s2"] = s2

    # S3 筹码形态 /20
    r_super = float(np.nanmean(super_net[:min(5, n)]))
    r_large = float(np.nanmean(large_net[:min(5, n)]))
    r_mid   = float(np.nanmean(mid_net[:min(5, n)]))
    r_small = float(np.nanmean(small_net[:min(5, n)]))
    classic_s = r_super > 0 and r_small < 0 and r_mid < 0
    classic   = r_super > 0 and r_small < 0
    retail    = r_super > 0 and r_small > 0 and r_small > abs(r_super) * 0.5
    quant     = abs(r_super) > 300 and abs(r_large) > 200 and np.sign(r_super) != np.sign(r_large)
    dump      = r_super < 0 and r_small > 0
    if   quant:                        s3 =  4; rk.append("超大单与大单方向相反，疑似量化对冲")
    elif dump:                         s3 =  2; rk.append("主力出货+散户接盘，风险极高")
    elif classic_s and recent_pct >= 2:s3 = 20; ev.append("超大单流入+中小单流出，经典机构建仓形态")
    elif classic   and recent_pct >= 2:s3 = 17; ev.append("超大单流入+小单流出，有效吸筹迹象")
    elif classic:                      s3 = 13; ev.append("超大单净流入同时小单净流出")
    elif retail:                       s3 =  8; rk.append("散户同步追涨，主力信号被稀释")
    elif r_super > 0:                  s3 = 11; ev.append("超大单净流入，筹码结构尚可")
    else:                              s3 =  6
    base["s3"] = s3

    # S4 趋势一致性 /10
    dirs    = np.sign(super_net[:n])
    dirs_nz = dirs[dirs != 0]
    if len(dirs_nz) == 0: s4 = 3
    else:
        cons = float(abs(dirs_nz.mean()))
        if   cons >= 0.85: s4 = 10; (ev if dirs_nz[0] > 0 else rk).append(f"近{n}日方向高度一致({cons:.0%})")
        elif cons >= 0.65: s4 =  7
        elif cons >= 0.45: s4 =  4; rk.append("近期资金流向方向多变")
        else:              s4 =  1; rk.append("资金流向混乱，量化对冲可能性高")
    base["s4"] = s4

    # S5 加速度 /15
    seq        = super_net[:n][::-1]
    slope      = 0.0
    if n >= 4:
        try: slope = float(np.polyfit(range(n), seq, 1)[0])
        except: pass
    mean_abs   = float(np.nanmean(np.abs(seq))) if n > 0 else 1
    slope_norm = slope / max(mean_abs, 1)
    if   slope_norm >=  0.30: s5 = 15; ev.append(f"资金净流入加速({slope_norm:+.2f}x)，主力提速建仓")
    elif slope_norm >=  0.12: s5 = 12; ev.append(f"资金净流入温和加速({slope_norm:+.2f}x)")
    elif slope_norm >=  0.03: s5 =  9; ev.append("资金净流入略有加速")
    elif slope_norm >= -0.03: s5 =  6
    elif slope_norm >= -0.15: s5 =  3; rk.append("资金净流入动能减弱")
    else:                     s5 =  0; rk.append(f"资金加速恶化({slope_norm:+.2f}x)，主力加速撤退")
    base["s5"] = s5

    # S6 价量背离 /10
    dark_days = dump_days = 0
    for i in range(min(5, n)):
        p, sv = pct_chg[i], super_net[i]
        if p <= -0.5 and sv > 0: dark_days += 1
        if p >=  2.0 and sv < 0: dump_days += 1
    if   dump_days >= 2: s6 =  0; rk.append(f"{dump_days}日阳线出货，拉高派发嫌疑")
    elif dump_days == 1: s6 =  3; rk.append("出现1日阳线出货，需警惕")
    elif dark_days >= 3: s6 = 10; ev.append(f"{dark_days}日阴线期间超大单持续流入，主力压价吸筹")
    elif dark_days == 2: s6 =  9; ev.append(f"{dark_days}日阴线吸筹，强建仓信号")
    elif dark_days == 1: s6 =  7; ev.append("出现阴线吸筹，主力压价建仓迹象")
    else:                s6 =  5
    base["s6"] = s6

    # N9 大单协同 /5
    large_corr = float(np.corrcoef(super_net[:n], large_net[:n])[0, 1]) if n >= 4 else 0.0
    if not quant and r_super > 0 and r_large > 0:
        if   large_corr >= 0.8: n9 = 5; ev.append(f"大单与超大单高度协同(r={large_corr:.2f})，机构共识极强")
        elif large_corr >= 0.5: n9 = 4; ev.append(f"大单与超大单同向协同(r={large_corr:.2f})")
        elif large_corr >= 0:   n9 = 3; ev.append("大单与超大单同向流入")
        else:                   n9 = 1
    elif quant or np.sign(r_super) != np.sign(r_large): n9 = 0
    else: n9 = 1
    base["n9_corr"] = n9

    # N12 资金效率比 /8
    pct5_sum = float(np.nansum(pct_chg[:min(5, n)]))
    if abs(pct5_sum) > 0.1 and net_5d != 0:
        eff      = net_5d / max(abs(pct5_sum), 0.1)
        eff_norm = eff / (circ_cap_yi * 100 if circ_cap_yi else 5000)
        if net_5d > 0:
            if   eff_norm >= 2.0: n12 = 8; ev.append("资金效率极高：大量净流入但涨幅有限，主力压价吸筹")
            elif eff_norm >= 0.8: n12 = 6; ev.append("资金效率较高：净流入与涨幅比例合理")
            elif eff_norm >= 0.2: n12 = 4
            else:                 n12 = 2; rk.append("资金效率低：资金推动涨幅过大，可能已透支")
        else: n12 = 1
    else: n12 = 3
    base["n12_eff"] = n12

    # N13 流入规律性 /6
    nonzero = super_net[super_net != 0]
    cv = np.nan
    if len(nonzero) >= 4:
        cv = float(np.std(nonzero) / (np.mean(np.abs(nonzero)) + 1e-6))
        if   cv < 0.5: n13 = 6; ev.append(f"资金流入极规律(CV={cv:.2f})，机构按计划建仓")
        elif cv < 1.0: n13 = 4; ev.append(f"资金流入有节奏(CV={cv:.2f})")
        elif cv < 2.0: n13 = 2
        else:          n13 = 0; rk.append(f"资金脉冲式(CV={cv:.2f})，游资特征")
    else: n13 = 2
    base["n13_cv"] = n13

    # N14 短中期加速比 /6
    mean_full  = float(np.nanmean(super_net[:n]))
    mean_short = float(np.nanmean(super_net[:min(3, n)]))
    st_ratio   = mean_short / (mean_full + 1e-6) if abs(mean_full) > 10 else 1.0
    if abs(mean_full) > 10:
        if   st_ratio >= 2.0: n14 = 6; ev.append(f"近3日资金加速({st_ratio:.1f}x全期均值)，主力提速")
        elif st_ratio >= 1.2: n14 = 5; ev.append(f"近期资金略有提速({st_ratio:.1f}x)")
        elif st_ratio >= 0.6: n14 = 3
        elif st_ratio >= 0:   n14 = 1; rk.append("近期资金明显减速，建仓动力衰减")
        else:                 n14 = 0; rk.append("近3日已转为净流出，主力近期撤退")
    else: n14 = 3
    base["n14_ratio"] = n14

    # ── 二期 K 线衍生指标 ─────────────────────────────────────
    if kdf is not None and len(kdf) >= 10:
        # N5 CMF /10
        cmf = calc_cmf(kdf, n=min(14, len(kdf)))
        base["cmf_val"] = cmf
        if not np.isnan(cmf):
            if   cmf >=  0.15: n5 = 10; ev.append(f"CMF={cmf:+.3f}，强积累买压")
            elif cmf >=  0.05: n5 =  8; ev.append(f"CMF={cmf:+.3f}，温和积累")
            elif cmf >=  0:    n5 =  5
            elif cmf > -0.05:  n5 =  3; rk.append(f"CMF={cmf:+.3f}，轻微分发压力")
            elif cmf > -0.15:  n5 =  1; rk.append(f"CMF={cmf:+.3f}，分发压力明显")
            else:              n5 =  0; rk.append(f"CMF={cmf:+.3f}，强分发，出货信号")
            if recent_pct > 1 and cmf > 0.05:  ev.append("超大单净占比+CMF双重确认，信号可靠")
            elif recent_pct > 1 and cmf < -0.05: rk.append("超大单净流入但CMF为负，信号矛盾")
        else: n5 = 4
        base["n5_cmf"] = n5

        # N7 OBV /8
        obv_s = calc_obv_slope(kdf, n=min(10, len(kdf) - 1))
        base["obv_slope"] = obv_s
        if not np.isnan(obv_s):
            if   obv_s >=  0.15: n7 = 8; ev.append(f"OBV快速上升({obv_s:+.2f})，量能强力确认趋势")
            elif obv_s >=  0.05: n7 = 6; ev.append("OBV平稳上升，趋势得到量能支撑")
            elif obv_s >=  0:    n7 = 4
            elif obv_s > -0.10:  n7 = 2; rk.append("OBV趋势下行，量能背离")
            else:                n7 = 0; rk.append(f"OBV快速下降({obv_s:+.2f})，出货/抛压信号")
        else: n7 = 3
        base["n7_obv"] = n7

        # N6 MFI /8
        mfi = calc_mfi(kdf, n=min(14, len(kdf) - 1))
        base["mfi_val"] = mfi
        if not np.isnan(mfi):
            if   mfi < 20:        n6 = 7; ev.append(f"MFI={mfi:.0f}，超卖区，逢低吸筹信号")
            elif 40 <= mfi <= 70: n6 = 6; ev.append(f"MFI={mfi:.0f}，资金健康流入区间")
            elif 70 < mfi <= 80:  n6 = 4
            elif mfi > 80:        n6 = 1; rk.append(f"MFI={mfi:.0f}，超买区，注意过热")
            else:                 n6 = 3
        else: n6 = 3
        base["n6_mfi"] = n6

        # N10 VWAP 偏离 /8
        vd = calc_vwap_dev(kdf, n=min(10, len(kdf)))
        base["vwap_dev"] = vd
        if not np.isnan(vd):
            if   vd < -2 and r_super > 0: n10 = 8; ev.append(f"收盘低于VWAP{abs(vd):.1f}%但超大单净流入，压价吸筹最强信号")
            elif vd <  0 and r_super > 0: n10 = 6; ev.append(f"收盘略低于VWAP({vd:+.1f}%)，主力均价下方吸筹")
            elif 0 <= vd <= 3:            n10 = 5
            elif 3 < vd <= 7:             n10 = 3
            else:                         n10 = 1; rk.append(f"收盘大幅高于VWAP({vd:+.1f}%)，短线过热")
        else: n10 = 3
        base["n10_vwap"] = n10

        # N11 量能形态 /8  +  N8 控盘度 /8
        vp = calc_vol_pattern(kdf)
        base["avg_turnover"] = vp["avg_to"]
        shrink, expand = vp["shrink_up"], vp["expand_down"]
        if   expand >= 3: n11 = 0; rk.append(f"近期{expand}日放量下跌，出货/抛压严重")
        elif expand >= 2: n11 = 2; rk.append(f"近期{expand}日放量下跌，需警惕")
        elif shrink >= 5: n11 = 8; ev.append(f"近期{shrink}日缩量上涨，筹码锁定良好")
        elif shrink >= 3: n11 = 6; ev.append(f"近期{shrink}日缩量上涨，浮筹减少")
        elif shrink >= 1: n11 = 4
        else:             n11 = 3
        base["n11_vol"] = n11

        n8 = vp["lock"]
        if n8 >= 7:  ev.append("近20日累计换手率极低，筹码高度锁定，控盘强")
        elif n8 >= 4: ev.append("筹码集中度较好")
        elif n8 <= 1: rk.append("换手频繁，筹码松散，主力控盘弱")
        base["n8_lock"] = n8

        # R4 高位风险
        hr = calc_high_risk(kdf)
        dist_h, dist_l = hr["dist_h"], hr["dist_l"]
        base["dist_high"] = dist_h
        if not np.isnan(dist_h):
            if   dist_h < 3:  r4 = -8; rk.append(f"距60日高点仅{dist_h:.1f}%，高位追入风险极大")
            elif dist_h < 8:  r4 = -5; rk.append(f"距60日高点{dist_h:.1f}%，接近前高，注意压力")
            elif dist_h < 15: r4 = -2
            else:             r4 =  0
        else: r4 = 0
        if not np.isnan(dist_l) and dist_l > 80:
            r4 = min(r4, -3); rk.append(f"距60日低点已反弹{dist_l:.0f}%，注意高位风险")
        base["r4_high"] = r4

        # R5 换手异常
        avg_to = vp["avg_to"]
        if not np.isnan(avg_to):
            if   avg_to > 15:        r5 = -5; rk.append(f"日均换手{avg_to:.1f}%过高，短线炒作特征")
            elif 3 <= avg_to <= 8:   r5 =  3; ev.append(f"日均换手{avg_to:.1f}%，机构稳健建仓节奏")
            elif avg_to < 0.8:       r5 = -3; rk.append(f"日均换手{avg_to:.1f}%过低，流动性不足")
            else: r5 = 0
        else: r5 = 0
        base["r5_turn"] = r5
    else:
        for k in ["n5_cmf", "n7_obv", "n6_mfi", "n10_vwap", "n11_vol", "n8_lock"]:
            base[k] = 3
        base["r4_high"] = base["r5_turn"] = 0

    # ── 奖惩系数 ──────────────────────────────────────────────
    coef    = 1.0
    abs_net5 = abs(net_5d)
    if circ_cap_yi:
        thr = 500 if circ_cap_yi < 50 else (3000 if circ_cap_yi < 200 else 10000)
        if abs_net5 < thr and r_super > 0:
            coef = min(coef, 0.85)
            rk.append(f"5日净额{abs_net5:.0f}万低于{circ_cap_yi:.0f}亿市值门槛{thr}万，信号可靠性存疑")
        elif abs_net5 >= thr * 3 and r_super > 0:
            base["bonus"] += 3
            ev.append(f"5日净额{abs_net5:.0f}万，远超门槛，资金量级充足")
    if ci == 1 and net_in_days == 1:
        coef = min(coef, 0.90)
        rk.append("超大单仅单日脉冲，短线游资可能性高")
    if dump_days >= 2:
        coef = min(coef, 0.80)
        rk.append(f"阳线出货×{dump_days}日，触发拉高派发惩罚")
    base["penalty_coef"] = coef

    # ── 汇总得分 ──────────────────────────────────────────────
    ff_raw = (s1 + s2 + s3 + s4 + s5 + s6 + n9 +
              base["n12_eff"] + base["n13_cv"] + base["n14_ratio"] +
              base["n5_cmf"] + base["n7_obv"] + base["n6_mfi"] + base["n10_vwap"] +
              base["n11_vol"] + base["n8_lock"] + base["r4_high"] + base["r5_turn"] + base["bonus"])
    total = max(0, min(100, int(round(ff_raw / 152 * 100 * coef))))
    base["total"] = total

    if   total >= 75: direction = "▲▲ 强势建仓"
    elif total >= 60: direction = "▲ 积极流入"
    elif total >= 45: direction = "→ 观望"
    elif total >= 30: direction = "↘ 谨慎"
    else:             direction = "▼ 疑似出货"
    base["direction"] = direction

    if   quant:                       pattern = "⚡ 量化对冲"
    elif dump_days >= 2:              pattern = "📦 拉高派发"
    elif dump and co >= 2:            pattern = "🚨 出货接盘"
    elif ci >= 4 and classic_s:       pattern = "🏗️  加速建仓"
    elif dark_days >= 2 and ci >= 2:  pattern = "🌑 压价吸筹"
    elif ci >= 3 and classic:         pattern = "📈 持续建仓"
    elif dark_days >= 2:              pattern = "🌑 阴线吸筹"
    elif ci >= 2 and classic:         pattern = "🔍 吸筹迹象"
    elif slope_norm >= 0.3 and ci >= 2: pattern = "🚀 加速启动"
    elif retail:                      pattern = "🌊 散户追涨"
    elif ci == 1 and net_in_days == 1: pattern = "⚡ 单日脉冲"
    elif co >= 3:                     pattern = "📦 持续出货"
    elif net_in_days > n // 2:        pattern = "🔄 间歇流入"
    else:                             pattern = "❓ 方向不明"
    base["pattern"] = pattern

    base.update({
        "consec_in": ci, "consec_out": co, "recent_pct": recent_pct,
        "net_3d": net_3d, "net_5d": net_5d, "slope_norm": slope_norm,
        "large_corr": large_corr, "dark_pool_days": dark_days, "dump_days": dump_days,
        "flow_cv": float(cv) if not np.isnan(cv) else 0.0,
        "st_ratio": st_ratio,
    })
    return base

# ══════════════════════════════════════════════════════════════
# 输出：控制台报告 + Excel
# ══════════════════════════════════════════════════════════════

def _bar(score, w=20):
    f = round(score / 100 * w)
    return f"[{'█'*f}{'░'*(w-f)}] {score}/100"

def _star(score):
    if score >= 75: return "★★★★★"
    if score >= 60: return "★★★★☆"
    if score >= 45: return "★★★☆☆"
    if score >= 30: return "★★☆☆☆"
    return "★☆☆☆☆"

def print_report(combined: list[dict]):
    print("\n" + "═" * 70)
    print("  资金选股策略 v4（Tushare 稳定版）—  综合分析报告")
    print(f"  {datetime.today().strftime('%Y-%m-%d')}  量价筛选 → 资金流向深度评分（18分项）")
    print("═" * 70)
    for rank, r in enumerate(combined, 1):
        s1_score = r.get("stage1_score", 0)
        s2_score = r.get("total", 0)
        has_flow = r.get("has_ff", False)
        cap      = r.get("circ_cap_yi", 0.0) or 0.0
        net3d_yi = r.get("net_3d", 0.0) / 1e4

        print(f"\n  {'─'*66}")
        print(f"  [{rank:>2}]  {r.get('name','')}（{r.get('code','')}）")
        to_str = f"{r.get('turnover', 0):.2f}%" if pd.notna(r.get("turnover")) else "─%"
        print(f"        现价：{r.get('price', 0.0):.2f}  "
              f"涨幅：{r.get('pct_chg', 0.0):+.2f}%  "
              f"量比：{r.get('vol_ratio', 0.0):.2f}x  "
              f"换手：{to_str}  "
              f"市值：{cap:.1f}亿")
        print(f"  ┌─ 第一阶段·量价评分 {'─'*40}")
        print(f"  │  得分：{s1_score}/8   命中：{' | '.join(r.get('stage1_hits', [])) or '─'}")
        print(f"  ├─ 第二阶段·资金流向深度评分（v3，满分100）{'─'*20}")
        if has_flow:
            print(f"  │  综合评分：{_bar(s2_score)}  {_star(s2_score)}")
            print(f"  │  方向预判：{r.get('direction','→ 观望')}    行为模式：{r.get('pattern','─')}")
            print(f"  │  分项：持续性{r.get('s1',0)}/25  强度{r.get('s2',0)}/20  "
                  f"筹码{r.get('s3',0)}/20  一致性{r.get('s4',0)}/10  "
                  f"加速{r.get('s5',0)}/15  背离{r.get('s6',0)}/10")
            print(f"  │       N9协同{r.get('n9_corr',0)}/5  CMF{r.get('n5_cmf',0)}/10  "
                  f"OBV{r.get('n7_obv',0)}/8  MFI{r.get('n6_mfi',0)}/8  "
                  f"VWAP{r.get('n10_vwap',0)}/8  量能{r.get('n11_vol',0)}/8  控盘{r.get('n8_lock',0)}/8")
            pct_s = "+" if r.get("recent_pct", 0.0) >= 0 else ""
            net_s = "+" if r.get("net_3d", 0.0) >= 0 else ""
            print(f"  │  超大单近3日占比：{pct_s}{r.get('recent_pct',0.0):.2f}%   "
                  f"近3日净流入：{net_s}{net3d_yi:.2f}亿   "
                  f"惩罚系数：{r.get('penalty_coef',1.0):.2f}")
            if r.get("evidence"):
                print("  │  ✦ 看涨依据：")
                for e in r["evidence"]:
                    print(f"  │    · {e}")
            if r.get("risk"):
                print("  │  ✦ 风险提示：")
                for ri in r["risk"]:
                    print(f"  │    · {ri}")
        else:
            print("  │  ⚠️ 未获取到资金流向数据")
        print(f"  └{'─'*62}")
    print(f"\n{'═'*70}")
    print("  ⚠️  仅供技术分析参考，不构成投资建议。投资有风险，入市须谨慎。")
    print(f"{'═'*70}\n")


def save_excel(combined: list[dict], out_path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = Workbook()

    def fill(h):  return PatternFill("solid", fgColor=h)
    def fnt(h, bold=False, sz=10): return Font(color=h, bold=bold, size=sz, name="Arial")
    def bd():
        t = Side(style="thin", color="D0D0D0")
        return Border(left=t, right=t, top=t, bottom=t)
    def ctr(): return Alignment(horizontal="center", vertical="center")
    def lft(wrap=False): return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

    HDR = fill("1A2B4A"); HF = fnt("FFFFFF", bold=True); NRM = fnt("333333")
    G = fill("C8F5D0"); T = fill("D0F0F8"); Y = fill("FFF9C4")
    O = fill("FFE0B2"); R = fill("FFCDD2"); BD = bd()

    def rfill(s):
        if s >= 75: return G
        if s >= 60: return T
        if s >= 45: return Y
        if s >= 30: return O
        return R

    def fmt(v, dec=2):
        if v is None or (isinstance(v, float) and np.isnan(v)): return "─"
        return round(v, dec)

    # ── Sheet 1 综合评分 ─────────────────────────────────────
    ws = wb.active
    ws.title = "综合评分"
    H1 = [
        "排名", "代码", "名称", "现价", "涨幅%", "量比", "换手%", "市值(亿)",
        "量价得分", "量价命中条件",
        "资金流向总分", "星级", "方向预判", "行为模式", "惩罚系数",
        # 资金流向一期
        "S1持续(25)", "S2强度(20)", "S3筹码(20)", "S4一致(10)", "S5加速(15)", "S6背离(10)",
        # 新增一期
        "N9协同(5)", "N12效率(8)", "N13规律(6)", "N14加速比(6)",
        # 二期
        "N5_CMF(10)", "N7_OBV(8)", "N6_MFI(8)", "N10_VWAP(8)", "N11量能(8)", "N8控盘(8)",
        "R4高位", "R5换手",
        # 关键数值
        "连续流入", "连续流出", "近3日占比%", "5日净额(万)",
        "加速斜率", "大单协同r", "阴线吸筹", "阳线出货",
        "CMF值", "OBV斜率", "MFI值", "VWAP偏离%", "距高点%", "日均换手%",
        "看涨依据", "风险提示",
    ]
    ws.append(H1)
    for cell in ws[1]:
        cell.fill = HDR; cell.font = HF; cell.alignment = ctr(); cell.border = BD
    ws.row_dimensions[1].height = 26

    for rank, r in enumerate(combined, 1):
        tot = r["total"]
        row = [
            rank, r["code"], r.get("name", r["code"]),
            fmt(r.get("price"), 2), fmt(r.get("pct_chg"), 2),
            fmt(r.get("vol_ratio"), 2),
            fmt(r["turnover"], 2) if pd.notna(r.get("turnover")) else "─",
            fmt(r.get("circ_cap_yi"), 1),
            r.get("stage1_score", 0),
            " | ".join(r.get("stage1_hits", [])) or "─",
            tot, _star(tot), r["direction"], r["pattern"],
            fmt(r.get("penalty_coef", 1.0), 2),
            r.get("s1",0), r.get("s2",0), r.get("s3",0), r.get("s4",0),
            r.get("s5",0), r.get("s6",0),
            r.get("n9_corr",0), r.get("n12_eff",0), r.get("n13_cv",0), r.get("n14_ratio",0),
            r.get("n5_cmf",0), r.get("n7_obv",0), r.get("n6_mfi",0),
            r.get("n10_vwap",0), r.get("n11_vol",0), r.get("n8_lock",0),
            r.get("r4_high",0), r.get("r5_turn",0),
            r.get("consec_in",0), r.get("consec_out",0),
            fmt(r.get("recent_pct"),2), fmt(r.get("net_5d"),0),
            fmt(r.get("slope_norm"),3), fmt(r.get("large_corr"),2),
            r.get("dark_pool_days",0), r.get("dump_days",0),
            fmt(r.get("cmf_val"),3), fmt(r.get("obv_slope"),3),
            fmt(r.get("mfi_val"),1), fmt(r.get("vwap_dev"),1),
            fmt(r.get("dist_high"),1), fmt(r.get("avg_turnover"),1),
            " | ".join(r.get("evidence", [])) or "─",
            " | ".join(r.get("risk", [])) or "─",
        ]
        ws.append(row)
        er = ws[ws.max_row]
        for cell in er:
            cell.fill = rfill(tot); cell.alignment = ctr(); cell.border = BD; cell.font = NRM
        er[10].font = fnt("1A2B4A", bold=True, sz=11)  # 总分
        er[-1].alignment = lft(wrap=True)
        er[-2].alignment = lft(wrap=True)
        er[9].alignment  = lft(wrap=True)
        if isinstance(er[14].value, (int, float)) and er[14].value < 1.0:
            er[14].font = fnt("B71C1C", bold=True)
        for idx in [32, 33]:   # R4/R5 负值红色
            v = er[idx].value
            if isinstance(v, (int, float)) and v < 0:
                er[idx].font = fnt("B71C1C", bold=True)

    cw = [5, 8, 12, 7, 7, 7, 7, 9,
          8, 36,
          8, 8, 14, 14, 8,
          10, 9, 9, 9, 9, 9,
          8, 8, 7, 9,
          8, 8, 7, 9, 8, 8, 7, 7,
          8, 8, 9, 11, 9, 9, 8, 8,
          8, 8, 7, 9, 8, 8,
          55, 50]
    for i, w in enumerate(cw, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.conditional_formatting.add(
        f"K2:K{len(combined)+1}",
        ColorScaleRule(start_type="min", start_color="FFCDD2",
                       mid_type="num", mid_value=50, mid_color="FFF9C4",
                       end_type="max", end_color="C8F5D0"))

    # ── Sheet 2 分项热力图 ────────────────────────────────────
    ws2 = wb.create_sheet("分项得分")
    h2_keys = [
        ("s1", 25), ("s2", 20), ("s3", 20), ("s4", 10), ("s5", 15), ("s6", 10),
        ("n9_corr", 5), ("n12_eff", 8), ("n13_cv", 6), ("n14_ratio", 6),
        ("n5_cmf", 10), ("n7_obv", 8), ("n6_mfi", 8),
        ("n10_vwap", 8), ("n11_vol", 8), ("n8_lock", 8),
        ("r4_high", 8), ("r5_turn", 5),
    ]
    h2_labels = ["代码", "名称"] + [f"{k}({mx})" for k, mx in h2_keys] + ["总分"]
    ws2.append(h2_labels)
    for cell in ws2[1]:
        cell.fill = HDR; cell.font = HF; cell.alignment = ctr(); cell.border = BD
    for r in combined:
        row2 = [r["code"], r.get("name", "─")] + [r.get(k, 0) for k, _ in h2_keys] + [r["total"]]
        ws2.append(row2)
        er2 = ws2[ws2.max_row]
        for cell in er2: cell.alignment = ctr(); cell.border = BD; cell.font = NRM
        er2[-1].fill = rfill(r["total"]); er2[-1].font = fnt("1A2B4A", bold=True)
        for ci2, (key, maxv) in enumerate(h2_keys, start=3):
            v = r.get(key, 0)
            if maxv <= 0: continue
            pct = v / maxv
            if   pct >= 0.75: er2[ci2 - 1].fill = G
            elif pct >= 0.50: er2[ci2 - 1].fill = T
            elif pct >= 0.25: er2[ci2 - 1].fill = Y
            else:             er2[ci2 - 1].fill = R
            if isinstance(v, (int, float)) and v < 0:
                er2[ci2 - 1].font = fnt("B71C1C", bold=True)
    for i, w in enumerate([8, 12] + [9] * len(h2_keys) + [8], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ── Sheet 3 评分说明 ──────────────────────────────────────
    ws3 = wb.create_sheet("评分说明")
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["C"].width = 30

    def sec(t):
        ws3.append([]); ws3.append([t])
        r = ws3[ws3.max_row]; r[0].font = Font(bold=True, size=12, color="1A2B4A", name="Arial")
        r[0].fill = fill("E3F2FD")

    def row3(a, b, c=""):
        ws3.append([a, b, c])
        r = ws3[ws3.max_row]; r[0].font = Font(bold=True, size=10, name="Arial")
        r[1].alignment = Alignment(wrap_text=True, vertical="center")
        ws3.row_dimensions[ws3.max_row].height = 26

    sec("v4 评分体系（三步走架构，资金流向满分100分）")
    row3("步骤1 全量快照",    "Tushare pro.daily + pro.daily_basic", "替代原东财爬虫")
    row3("步骤2 量价评分",    "8条条件，本地无请求，K线来自 Tushare pro_bar", "")
    row3("步骤3 资金流向评分","18分项，满分100，来自 fund_flow_scorer_v6 完整引擎", "")
    row3("惩罚系数",          "乘法叠加：金额门槛×0.85 / 单日脉冲×0.90 / 阳线出货×0.80", "")

    sec("量价评分条件（步骤2，满分8分）")
    row3("① 涨幅 3-8%",       "绝对涨幅区间筛选", "")
    row3("② 换手 5-15%",      "有效换手区间", "")
    row3("③ 放量≥1.5x",       "当日量/前5日均量", "")
    row3("④ 超额+1%",         "跑赢沪深300超过1%", "")
    row3("⑤ 量TOP10%",        "近60日量能百分位", "")
    row3("⑥ 价格>MA13>MA34>MA55", "均线多头排列（斐波那契均线）", "")
    row3("⑦ 近5日放量阳线≥2根","量价齐升确认", "")
    row3("⑧ 距高点回调<15%",  "回调幅度合理", "")

    sec("资金流向各分项（步骤3，原始满分约152→压缩到100）")
    row3("S1 持续性 /25",     "超大单连续净流入天数", "≥5日=25，4=22，3=18，2=13，1=7")
    row3("S2 强度 /20",       "近3日超大单净占比", "≥6%=20，≥4%=17，≥2%=13，≥0.8%=8")
    row3("S3 筹码形态 /20",   "超大/中/小单方向组合", "经典建仓=20，量化对冲=4，出货=2")
    row3("S4 趋势一致性 /10", "N日方向稳定度", "≥85%=10，≥65%=7，≥45%=4")
    row3("S5 加速度 /15",     "净额归一化线性斜率", "≥0.3x=15，≥0.12x=12，≥0.03x=9")
    row3("S6 价量背离 /10",   "阴线吸筹/阳线出货", "阴线吸筹≥3日=10，阳线出货≥2=0")
    row3("N9 大单协同 /5",    "大单与超大单相关系数", "r≥0.8=5，r≥0.5=4，同向=3")
    row3("N12 资金效率比 /8", "5日净额/累计涨幅", "效率高=主力压价建仓")
    row3("N13 流入规律性 /6", "变异系数CV", "CV<0.5=6（机构），CV>2=0（游资）")
    row3("N14 短中期加速比 /6","近3日/全期均值", ">2x=6，>1.2x=5")
    row3("N5 CMF /10",        "蔡金资金流", ">0.15=10，<-0.15=0")
    row3("N7 OBV /8",         "能量潮斜率", "上升=8，背离预警")
    row3("N6 MFI /8",         "资金流量指数", "20-70健康区，>80超买")
    row3("N10 VWAP偏离 /8",   "收盘价vs成交均价", "低于VWAP+净流入=压价吸筹")
    row3("N11 量能形态 /8",   "缩量上涨/放量下跌", "缩量上涨≥5日=8，放量下跌≥3=0")
    row3("N8 控盘度 /8",      "近20日换手率", "累计<15%=极高控盘")
    row3("R4 高位风险 0~-8",  "距60日最高价", "<3%=-8，<8%=-5")
    row3("R5 换手异常 ±5",    "日均换手率", "3-8%=+3（理想），>15%=-5（过热）")

    sec("评级标准")
    row3("≥75分 ★★★★★",  "强势建仓", "资金信号极强，主力持续重仓布局")
    row3("60-74分 ★★★★☆", "积极流入", "主力积极布局，可重点跟踪")
    row3("45-59分 ★★★☆☆", "观望",     "信号偏弱，等待更强确认")
    row3("30-44分 ★★☆☆☆", "谨慎",     "资金分歧，不宜追高")
    row3("<30分 ★☆☆☆☆",  "疑似出货", "主力撤退风险高，需回避")

    ws3.append([])
    ws3.append(["⚠️ 免责声明", "本程序仅供技术分析参考，不构成投资建议。投资有风险，入市须谨慎。"])
    wb.save(out_path)
    if isinstance(out_path, (str, pathlib.Path)):
        print(f"  ✅ Excel 已保存：{out_path}")
    else:
        print("  ✅ Excel 已生成（内存）")
